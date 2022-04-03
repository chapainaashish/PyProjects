
# Contact Manager using python tkinter and openpyxl
from tkinter import Tk, Entry, Label, Button
import openpyxl


# In this case we have to manually create *.xlsx file in project directory
# REASON: the file created from python file context manager will have empty metadata
# with no cell formatting which will raise 'Badzipfile' error
# with open("contact_book.xlsx", "a"):
#     pass

# loading workbook
mybook = openpyxl.load_workbook("/home/aashish/Documents/Python-Journey/Python-Projects/Contact_Manager/contact_book.xlsx")
sheet = mybook.active


def load_workbook(sheet, mybook):
    """Adding different field.get() if there doesn't exist any"""
    cellassign = sheet.cell(1, 1)
    if cellassign.value != "Name":
        sheet['A1'] = "Name"
        sheet['B1'] = "Phone"
        sheet['C1'] = "Email"
        sheet['D1'] = "Address"
        sheet['E1'] = "Website"
        mybook.save("contact_book.xlsx")


def check_write_data(*args):
    """Checking if user input match with database"""
    row = sheet.max_row
    for i in range(1, row + 1):
        name_obj = sheet.cell(i, 1)
        phone_obj = sheet.cell(i, 2)

        if name_field.get() == "" or phone_field.get() == "":
            warning_text = Label(root, text="Please, Enter Valid Input!",
                                 background="black", fg="red", font=("garuda", 11, "bold"))
            warning_text.grid(row=11, column=1)
            return False

        # Raising error if there exist any simialr contact
        if (name_obj.value == name_field.get()) or (phone_obj.value == phone_field.get()):
            warning_text = Label(root, text="Contact Already Exist\nChange your Name or Phone number and Try Again!",
                                 background="black", fg="red", font=("garuda", 11, "bold"))
            warning_text.grid(row=11, column=1)

            return False

    # Writing data if there doesn't exist any duplication
    else:
        data = ((name_field.get(), phone_field.get(), email_field.get(),
                address_field.get(), website_field.get()),)
        for row in data:
            sheet.append(row)
            mybook.save("contact_book.xlsx")

        warning_text = Label(root, text="Contact Successfully Saved",
                             background="black", fg="green", font=("garuda", 11, "bold"))
        warning_text.grid(row=11, column=1, ipadx=100, ipady=50)
        return True


# UI development
# Customizing root window
root = Tk()
root.geometry("588x433")
root.minsize(588, 433)
root.maxsize(688, 533)
root.configure(background="black")
root.title("Contact Manager")

# Adding header
root_text = Label(root, text="Contacts Saver", background="black",
                  fg="white", font=("garuda", 18, "bold"))
root_text.grid(column=1)

# Intializing Label
# Name, Phone, Email, Address, Website;
name = Label(root, text="Name", background="black", foreground="white")
phone = Label(root, text="Phone", background="black", foreground="white")
email = Label(root, text="Email", background="black", foreground="white")
address = Label(root, text="Address", background="black", foreground="white")
website = Label(root, text="Website", background="black", foreground="white")

# Packing Label as Grid
name.grid(row=5, column=0, ipadx=40)
phone.grid(row=6, column=0)
email.grid(row=7, column=0)
address.grid(row=8, column=0)
website.grid(row=9, column=0)

# Creating Entry Field
name_field = Entry(root, background="black", foreground="white")
phone_field = Entry(root, background="black", foreground="white")
email_field = Entry(root, background="black", foreground="white")
address_field = Entry(root, background="black", foreground="white")
website_field = Entry(root, background="black", foreground="white")

# Packing Entry field as grid
name_field.grid(row=5, column=1, pady=3, ipadx=100)
phone_field.grid(row=6, column=1, pady=3, ipadx=100)
email_field.grid(row=7, column=1, pady=3, ipadx=100)
address_field.grid(row=8, column=1, pady=3, ipadx=100)
website_field.grid(row=9, column=1, pady=3, ipadx=100)

# Creating Button
load_workbook(sheet, mybook)
submit_button = Button(root, text="Save", background="black",
                       foreground="white", command=check_write_data)
submit_button.grid(row=10, column=1, pady=15, ipadx=10)

# Focusing to  next entry field if "Enter" is pressed/"<Return>"
# But first initializing input to name entry field
name_field.bind("<Return>", func=name_field.focus_set())
name_field.bind("<Return>", lambda event: phone_field.focus_set())
phone_field.bind("<Return>", lambda event: email_field.focus_set())
email_field.bind("<Return>", lambda event: address_field.focus_set())
address_field.bind("<Return>", lambda event: website_field.focus_set())
website_field.bind("<Return>", lambda event: submit_button.focus_set())
submit_button.bind("<Return>", func=check_write_data)

root.mainloop()
